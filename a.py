from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── palette ───────────────────────────────────────────────────────────────────
DARK="1A1A2E"; TEAL="00897B"; PURPLE="5E35B1"
CUS="D4EDDA"; CAR="FFF3E0"; CZH="FFF8E1"; CEN="E3F2FD"
CFR="F3E5F5"; CDE="E8EAF6"; COT="FBE9E7"; CMX="FFCCBC"
WHITE="FFFFFF"; LGRAY="F8F9FA"

def tb():
    s=Side(style="thin",color="CCCCCC")
    return Border(left=s,right=s,top=s,bottom=s)

def sc(cell,bg=None,fc="000000",sz=10,bold=False,wrap=True,ha="left",va="top"):
    if bg: cell.fill=PatternFill("solid",fgColor=bg)
    cell.font=Font(name="Arial",size=sz,bold=bold,color=fc)
    cell.alignment=Alignment(horizontal=ha,vertical=va,wrap_text=wrap)
    cell.border=tb()

# ══════════════════════════════════════════════════════════════════════════════
#  ALL CONTENT
# ══════════════════════════════════════════════════════════════════════════════

WN_EN="""WELCOME TO VITAFLOW — Version 1.0

VitaFlow is now live. Everything included at launch:

• Blood pressure monitor with clinical colour-coding and trend charts
• Camera heart rate using optical PPG — no extra device needed
• Blood glucose tracker with meal-context tagging
• Blood oxygen (SpO2) tracking
• Step counter with live arc gauge and daily goal
• AI prescription scanner — photo your prescription, auto-fill your medication list
• Medication reminders with custom schedule, snooze, and repeat
• Weekly and monthly health charts for every vital
• PDF and CSV export — share your history with your doctor
• Apple Watch companion app
• Home screen and lock screen widgets
• iCloud backup and one-tap restore
• Dark mode, five accent colours, full personalisation

All data stays on your device — no account, no server, no tracking.

If you find VitaFlow useful, a review helps us reach more people who need it. Thank you for downloading."""

WN_AR="""مرحباً بكم في VitaFlow — الإصدار 1.0

VitaFlow متاح الآن. إليكم كل ما يتضمنه الإصدار الأول:

• مراقب ضغط الدم مع ترميز لوني سريري ومخططات اتجاه
• قياس معدل ضربات القلب عبر الكاميرا باستخدام تقنية PPG البصرية — لا حاجة لأجهزة إضافية
• متتبع جلوكوز الدم مع تحديد سياق الوجبة
• تتبع تشبع الأكسجين في الدم SpO2
• عداد الخطوات مع مقياس قوس مباشر وهدف يومي
• ماسح وصفات طبية بالذكاء الاصطناعي — صوّر وصفتك الطبية وأدخل قائمة أدويتك تلقائياً
• تذكير بالأدوية مع جدول مخصص وإمكانية التأجيل
• مخططات صحية أسبوعية وشهرية لكل مؤشر حيوي
• تصدير بصيغة PDF و CSV — شارك سجلاتك مع طبيبك
• تطبيق مرافق لـ Apple Watch
• ودجت الشاشة الرئيسية وشاشة القفل
• نسخ احتياطي على iCloud واستعادة بنقرة واحدة
• الوضع الليلي وخمسة ألوان مميزة وتخصيص كامل

جميع البيانات تبقى على جهازك — لا حساب، لا خادم، لا مشاركة بيانات.
إذا وجدت VitaFlow مفيداً، فإن التقييم يساعدنا في الوصول إلى المزيد. شكراً للتنزيل."""

WN_ZH_HANS="""欢迎使用 VitaFlow — 版本 1.0

VitaFlow 现已上线。首发版本包含的全部功能：

• 血压监测仪，配有临床色码和趋势图
• 摄像头心率检测（光学PPG技术）—— 无需额外设备
• 血糖追踪器，支持餐前/餐后/空腹标记
• 血氧饱和度（SpO2）监测
• 计步器，配有实时弧形仪表和每日目标
• AI处方扫描仪 —— 拍摄处方照片，自动填充药物清单
• 药物提醒，支持自定义时间表和贪睡功能
• 每项健康指标的周/月趋势图
• PDF和CSV导出 —— 与医生分享健康记录
• Apple Watch伴侣应用
• 主屏幕和锁定屏幕小组件
• iCloud备份和一键恢复
• 深色模式、五种强调色、完整个性化设置

所有数据保存在您的设备上 —— 无账户、无服务器、无数据共享。
如果您觉得VitaFlow有用，请留下评价。感谢您的下载。"""

WN_ZH_HANT="""歡迎使用 VitaFlow — 版本 1.0

VitaFlow 現已上線。首發版本包含的全部功能：

• 血壓監測儀，配有臨床色碼和趨勢圖
• 攝影機心率偵測（光學PPG技術）—— 無需額外裝置
• 血糖追蹤器，支援餐前/餐後/空腹標記
• 血氧飽和度（SpO2）監測
• 計步器，配有即時弧形儀表和每日目標
• AI處方掃描器 —— 拍攝處方照片，自動填入藥物清單
• 藥物提醒，支援自訂時間表和貪睡功能
• 每項健康指標的週/月趨勢圖
• PDF和CSV匯出 —— 與醫師分享健康記錄
• Apple Watch伴侣應用程式
• 主畫面和鎖定畫面小工具
• iCloud備份和一鍵還原
• 深色模式、五種強調色、完整個人化設定

所有資料保存在您的裝置上 —— 無帳戶、無伺服器、無資料共享。
如果您覺得VitaFlow有用，請留下評價。感謝您的下載。"""

WN_FR="""BIENVENUE SUR VITAFLOW — Version 1.0

VitaFlow est maintenant disponible. Voici tout ce qui est inclus :

• Moniteur de pression artérielle avec codage couleur clinique et graphiques
• Mesure de la fréquence cardiaque par caméra (PPG optique) — aucun appareil supplémentaire
• Suivi de la glycémie avec marquage du contexte repas
• Surveillance de la saturation en oxygène (SpO2)
• Podomètre avec jauge à arc en temps réel et objectif quotidien
• Scanner d'ordonnances IA — photographiez, remplissez automatiquement votre liste
• Rappels de médicaments avec planning personnalisé et rappel différé
• Graphiques de tendances hebdomadaires et mensuels pour chaque signe vital
• Export PDF et CSV — partagez vos données avec votre médecin
• Application compagnon Apple Watch
• Widgets écran d'accueil et écran verrouillé
• Sauvegarde iCloud et restauration en un appui
• Mode sombre, cinq couleurs d'accent, personnalisation complète

Toutes les données restent sur votre appareil — aucun compte, aucun serveur.
Si vous trouvez VitaFlow utile, un avis nous aide beaucoup. Merci."""

WN_DE="""WILLKOMMEN BEI VITAFLOW — Version 1.0

VitaFlow ist jetzt verfügbar. Das ist alles in der ersten Version:

• Blutdruckmessgerät mit klinischer Farbkodierung und Trenddiagrammen
• Herzfrequenzmessung per Kamera (optische PPG-Technologie) — kein Zusatzgerät
• Blutzucker-Tracker mit Mahlzeit-Kontext-Markierung
• Überwachung der Blutsauerstoffsättigung (SpO2)
• Schrittzähler mit Echtzeit-Bogenanzeige und Tagesziel
• KI-Rezeptscanner — fotografieren Sie Ihr Rezept, füllen Sie die Liste automatisch aus
• Medikamentenerinnerungen mit anpassbarem Zeitplan und Schlummerfunktion
• Wöchentliche und monatliche Trenddiagramme für jeden Vitalwert
• PDF- und CSV-Export — Teilen Sie Ihre Daten mit Ihrem Arzt
• Apple Watch Begleit-App
• Home Screen und Sperrbildschirm Widgets
• iCloud-Backup und Wiederherstellung per Tastendruck
• Dunkelmodus, fünf Akzentfarben, vollständige Personalisierung

Alle Daten bleiben auf Ihrem Gerät — kein Konto, kein Server.
Wenn Sie VitaFlow nützlich finden, hilft eine Bewertung sehr. Danke."""

WN_JA="""VitaFlowへようこそ — バージョン 1.0

VitaFlowがリリースされました。初版に含まれるすべての機能：

• 血圧モニター（臨床カラーコードとトレンドチャート付き）
• カメラによる心拍数測定（光学PPG技術）— 追加デバイス不要
• 血糖値トラッカー（食事コンテキストタグ付き）
• 血中酸素飽和度（SpO2）追跡
• 歩数計（リアルタイムアークゲージと日次目標付き）
• AI処方箋スキャナー — 処方箋を撮影するだけで薬リストを自動入力
• 薬の服用リマインダー（カスタムスケジュールとスヌーズ機能付き）
• 各バイタルの週次・月次トレンドチャート
• PDFおよびCSVエクスポート — 医師と記録を共有
• Apple Watchコンパニオンアプリ
• ホーム画面とロック画面のウィジェット
• iCloudバックアップとワンタップ復元
• ダークモード、5つのアクセントカラー、完全なカスタマイズ

すべてのデータはデバイスに保存されます — アカウント不要、サーバー不要。
VitaFlowが役立つと感じたら、ぜひレビューをお願いします。"""

WN_KO="""VitaFlow에 오신 것을 환영합니다 — 버전 1.0

VitaFlow이 출시되었습니다. 첫 번째 버전에 포함된 모든 기능:

• 임상 색상 코딩 및 추세 차트가 있는 혈압 모니터
• 카메라를 이용한 심박수 측정 (광학 PPG 기술) — 추가 기기 불필요
• 식사 맥락 태그가 있는 혈당 추적기
• 혈중 산소 포화도 (SpO2) 모니터링
• 실시간 아크 게이지와 일일 목표가 있는 만보기
• AI 처방전 스캐너 — 처방전을 촬영하면 약물 목록이 자동으로 채워짐
• 사용자 지정 일정과 다시 알림이 있는 약 복용 알림
• 각 활력징후의 주간 및 월간 추세 차트
• PDF 및 CSV 내보내기 — 의사에게 기록 공유
• Apple Watch 동반 앱
• 홈 화면 및 잠금 화면 위젯
• iCloud 백업 및 원터치 복원
• 다크 모드, 5가지 강조 색상, 완전한 개인화

모든 데이터는 기기에 저장됩니다 — 계정 없음, 서버 없음.
VitaFlow이 유용하다면 리뷰를 남겨주세요. 감사합니다."""

WN_PT="""BEM-VINDO AO VITAFLOW — Versão 1.0

O VitaFlow está disponível. Tudo incluído nesta primeira versão:

• Monitor de pressão arterial com codificação de cores clínica e gráficos de tendência
• Medição de frequência cardíaca por câmera (tecnologia PPG óptica) — sem dispositivo adicional
• Rastreador de glicose no sangue com marcação de contexto de refeição
• Monitoramento de saturação de oxigênio no sangue (SpO2)
• Contador de passos com medidor de arco em tempo real e meta diária
• Scanner de receitas com IA — fotografe sua receita, preencha automaticamente sua lista
• Lembretes de medicação com agenda personalizada e soneca
• Gráficos de tendência semanal e mensal para cada sinal vital
• Exportação em PDF e CSV — compartilhe seu histórico com seu médico
• Aplicativo companheiro para Apple Watch
• Widgets na tela inicial e na tela de bloqueio
• Backup no iCloud e restauração com um toque
• Modo escuro, cinco cores de destaque, personalização completa

Todos os dados ficam no seu dispositivo — sem conta, sem servidor.
Se você achar o VitaFlow útil, uma avaliação nos ajuda muito. Obrigado."""

WN_ES="""BIENVENIDO A VITAFLOW — Versión 1.0

VitaFlow ya está disponible. Todo lo incluido en esta primera versión:

• Monitor de presión arterial con codificación de colores clínica y gráficos de tendencia
• Medición de frecuencia cardíaca por cámara (tecnología PPG óptica) — sin dispositivos adicionales
• Rastreador de glucosa en sangre con etiquetado de contexto de comida
• Monitoreo de saturación de oxígeno en sangre (SpO2)
• Podómetro con medidor de arco en tiempo real y objetivo diario
• Escáner de recetas con IA — fotografía tu receta, completa automáticamente tu lista
• Recordatorios de medicación con horario personalizado y posponer
• Gráficos de tendencia semanal y mensual para cada signo vital
• Exportación en PDF y CSV — comparte tu historial con tu médico
• Aplicación complementaria para Apple Watch
• Widgets en pantalla de inicio y pantalla de bloqueo
• Copia de seguridad en iCloud y restauración con un toque
• Modo oscuro, cinco colores de acento, personalización completa

Todos los datos permanecen en tu dispositivo — sin cuenta, sin servidor.
Si encuentras útil VitaFlow, una reseña nos ayuda mucho. Gracias."""

# ── Descriptions ───────────────────────────────────────────────────────────────
DESC_EN="""ALL YOUR HEALTH VITALS IN ONE PLACE

VitaFlow is the most complete AI health tracker for iPhone. Monitor blood pressure, heart rate, blood glucose, blood oxygen (SpO2), and daily steps — all in one app. Powered by AI. Backed by clinical science. Your health data stays on your device.

KEY FEATURES

BLOOD PRESSURE MONITOR
Log systolic, diastolic, and pulse every reading. Colour-coded against clinical thresholds — green for normal, amber for elevated, red for high. Weekly and monthly trend charts. Share with your doctor in one tap.

AI PRESCRIPTION SCANNER
The health tracking feature no other app has. Point your camera at any prescription or medication bottle. VitaFlow AI reads the text and automatically populates your medication list with drug names, dosages, and schedules. No manual entry. Powered by Google Gemini.

CAMERA HEART RATE MONITOR
No device needed. Place your fingertip over the rear camera and flash. VitaFlow uses optical PPG technology to detect blood-volume pulse and measure your heart rate in real time.

BLOOD GLUCOSE TRACKER
Log blood glucose readings in mg/dL or mmol/L. Tag each reading with meal context — fasting, before meal, or after meal. Clinical colour-coded thresholds for diabetic and pre-diabetic ranges.

BLOOD OXYGEN (SpO2) MONITOR
Track blood oxygen saturation over time. Alerts when saturation drops below normal clinical range.

STEP COUNTER PEDOMETER
Live step counter with a custom arc gauge. Set your own daily goal. Weekly bar chart for the last 7 days. Full Apple Health integration.

MEDICATION REMINDER APP
Never miss a dose. Custom schedule per medication. Snooze for 1 hour. Daily and weekly repeat. Works fully offline.

HEALTH CHARTS AND HISTORY
Full reading history for every vital with swipe-to-delete. Weekly and monthly trend line charts for every vital type.

PDF AND CSV EXPORT
Export any vital history as formatted PDF or CSV. Share via email, AirDrop, or any share sheet.

APPLE WATCH COMPANION
Log readings from your wrist. Watch app displays latest vitals and syncs back to iPhone automatically. HealthKit integration included.

HOME SCREEN WIDGET
See all 5 vitals from your Home Screen or Lock Screen without opening the app. Small and medium WidgetKit sizes. Updates immediately every time you log a reading.

ICLOUD BACKUP
All data backed up to your personal iCloud. Restore on a new device in one tap. Merge-safe — never lose data.

PRIVACY FIRST
All readings stored locally on your device and in your personal iCloud. No external server. No account registration. No data sharing.

SUBSCRIPTION
VitaFlow is free to start with full core tracking. Premium unlocks advanced AI features, unlimited export, and Apple Watch sync.
Payment charged to iTunes Account at purchase confirmation. Auto-renews unless cancelled 24+ hours before period end. Manage in Apple ID Account Settings.

Privacy: https://sinceregen-legal.blogspot.com/p/privacy-policy.html
Terms: https://sinceregen-legal.blogspot.com/p/terms-conditions.html
Support: sinceregen.tech@gmail.com"""

DESC_AR="""جميع مؤشراتك الحيوية في مكان واحد

VitaFlow هو متتبع صحي AI الأكثر شمولاً لـ iPhone. راقب ضغط الدم ومعدل ضربات القلب وجلوكوز الدم وتشبع الأكسجين (SpO2) والخطوات اليومية — كل ذلك في تطبيق واحد.

الميزات الرئيسية

مراقب ضغط الدم
سجّل الانقباضي والانبساطي والنبض في كل قياس. ترميز لوني وفق المعايير السريرية. مخططات أسبوعية وشهرية.

ماسح الوصفات الطبية بالذكاء الاصطناعي
وجّه كاميرتك نحو أي وصفة طبية. يقرأ الذكاء الاصطناعي النص ويملأ قائمة أدويتك تلقائياً. لا إدخال يدوي. لا يوجد متتبع صحي آخر يمتلك هذه الميزة.

قياس معدل ضربات القلب بالكاميرا
ضع إصبعك على الكاميرا الخلفية. تقنية PPG البصرية تقيس معدل ضربات قلبك في ثوانٍ. لا حاجة لأجهزة إضافية.

متتبع جلوكوز الدم
التسجيل بـ mg/dL أو mmol/L. تحديد سياق الوجبة. تحديد لوني سريري.

مراقب الأكسجين في الدم (SpO2)
تتبع تشبع الأكسجين مع مرور الوقت.

عداد الخطوات
عداد خطوات مباشر مع مقياس قوس. هدف يومي قابل للتخصيص. تكامل كامل مع Apple Health.

تذكير بالأدوية
لا تفوّت جرعة. جدول مخصص لكل دواء. تأجيل ساعة. يعمل بدون إنترنت.

تصدير PDF و CSV
تصدير أي سجل صحي وإرساله مباشرة.

مرافق Apple Watch
تسجيل القراءات من معصمك. مزامنة تلقائية مع iPhone.

ودجت الشاشة الرئيسية
جميع المؤشرات الخمسة من شاشتك الرئيسية دون فتح التطبيق.

iCloud
نسخ احتياطي لجميع البيانات على iCloud الشخصي. الاستعادة بنقرة واحدة.

الخصوصية أولاً
جميع البيانات مخزنة محلياً. لا خادم خارجي. لا تسجيل حساب. لا مشاركة بيانات.

الخصوصية: https://sinceregen-legal.blogspot.com/p/privacy-policy.html
الشروط: https://sinceregen-legal.blogspot.com/p/terms-conditions.html
الدعم: sinceregen.tech@gmail.com"""

DESC_ZH_HANS="""所有健康数据，尽在一处

VitaFlow 是 iPhone 上功能最全面的 AI 健康追踪器。监测血压、心率、血糖、血氧饱和度（SpO2）和每日步数——一个应用搞定一切。

核心功能

血压监测仪
记录收缩压、舒张压和脉搏。根据临床阈值进行颜色编码。周/月趋势图。

AI 处方扫描仪
将摄像头对准任何处方。AI 读取文字并自动填充药物名称、剂量和时间表。无需手动输入。其他健康追踪器没有此功能。

摄像头心率测量
将指尖放在后置摄像头上，光学 PPG 技术在几秒内测量心率。无需额外设备。

血糖追踪器
支持 mg/dL 和 mmol/L 两种单位。餐前/餐后/空腹标记。临床阈值色码。

血氧饱和度（SpO2）监测
追踪血氧饱和度变化趋势。

计步器
实时计步器配弧形仪表。可配置每日目标。完整的 Apple Health 集成。

药物提醒
自定义每种药物的提醒时间表。贪睡一小时。完全离线工作。

健康图表和历史记录
每种健康指标的时间顺序记录和周/月趋势折线图。

PDF 和 CSV 导出
将任何健康记录导出并直接从应用分享。

Apple Watch 伴侣
从手腕记录数据。自动同步到 iPhone。

主屏幕小组件
无需打开应用，直接在主屏幕查看所有 5 项健康数据。

iCloud 备份
所有数据备份到个人 iCloud 账户。一键在新设备上恢复。

隐私优先
所有数据本地存储。无外部服务器，无账户注册，无数据共享。

隐私政策: https://sinceregen-legal.blogspot.com/p/privacy-policy.html
服务条款: https://sinceregen-legal.blogspot.com/p/terms-conditions.html
支持: sinceregen.tech@gmail.com"""

DESC_ZH_HANT="""所有健康數據，盡在一處

VitaFlow 是 iPhone 上功能最全面的 AI 健康追蹤器。監測血壓、心率、血糖、血氧飽和度（SpO2）和每日步數——一個應用程式搞定一切。

核心功能

血壓監測儀
記錄收縮壓、舒張壓和脈搏。根據臨床閾值進行顏色編碼。週/月趨勢圖。

AI 處方掃描器
將攝影機對準任何處方。AI 讀取文字並自動填入藥物名稱、劑量和時間表。無需手動輸入。其他健康追蹤器沒有此功能。

攝影機心率測量
將指尖放在後置攝影機上，光學 PPG 技術在幾秒內測量心率。無需額外裝置。

血糖追蹤器
支援 mg/dL 和 mmol/L 兩種單位。餐前/餐後/空腹標記。臨床閾值色碼。

血氧飽和度（SpO2）監測
追蹤血氧飽和度變化趨勢。

計步器
即時計步器配弧形儀表。可設定每日目標。完整的 Apple Health 整合。

藥物提醒
自訂每種藥物的提醒時間表。貪睡一小時。完全離線工作。

健康圖表和歷史記錄
每種健康指標的時間順序記錄和週/月趨勢折線圖。

PDF 和 CSV 匯出
將任何健康記錄匯出並直接從應用程式分享。

Apple Watch 伴侶
從手腕記錄資料。自動同步到 iPhone。

主畫面小工具
無需開啟應用，直接在主畫面查看所有 5 項健康數據。

iCloud 備份
所有資料備份到個人 iCloud 帳戶。一鍵在新裝置上還原。

隱私優先
所有資料本地儲存。無外部伺服器，無帳戶註冊，無資料共享。

隱私政策: https://sinceregen-legal.blogspot.com/p/privacy-policy.html
使用條款: https://sinceregen-legal.blogspot.com/p/terms-conditions.html
支援: sinceregen.tech@gmail.com"""

DESC_FR="""TOUS VOS SIGNES VITAUX EN UN SEUL ENDROIT

VitaFlow est le tracker de santé IA le plus complet pour iPhone. Surveillez la pression artérielle, la fréquence cardiaque, la glycémie, la saturation en oxygène (SpO2) et vos pas quotidiens — dans une seule application.

FONCTIONNALITÉS CLÉS

MONITEUR DE PRESSION ARTÉRIELLE
Systolique, diastolique et pouls. Codage couleur clinique. Graphiques hebdomadaires et mensuels.

SCANNER D'ORDONNANCES IA
Pointez votre caméra sur une ordonnance. L'IA remplit automatiquement votre liste de médicaments. Aucune saisie manuelle. Aucun autre tracker de santé n'a cette fonctionnalité.

MONITEUR DE FRÉQUENCE CARDIAQUE PAR CAMÉRA
Posez votre doigt sur la caméra arrière. PPG optique mesure la fréquence cardiaque en secondes. Aucun appareil supplémentaire.

TRACKER DE GLYCÉMIE
mg/dL ou mmol/L. Marquage du contexte repas. Seuils cliniques codés en couleur.

MONITEUR D'OXYGÈNE (SpO2)
Suivi de la saturation en oxygène dans le temps.

PODOMÈTRE
Compteur en temps réel. Objectif quotidien configurable. Intégration Apple Health.

RAPPELS DE MÉDICAMENTS
Planning personnalisé. Rappel différé. Fonctionne hors ligne.

EXPORT PDF ET CSV
Partagez directement depuis l'application.

APPLE WATCH
Enregistrez depuis le poignet. Synchronisation automatique avec iPhone.

ICLOUD
Sauvegarde dans votre compte iCloud personnel. Restauration en un appui.

CONFIDENTIALITÉ
Toutes les données stockées localement. Aucun serveur. Aucun compte.

Confidentialité: https://sinceregen-legal.blogspot.com/p/privacy-policy.html
Conditions: https://sinceregen-legal.blogspot.com/p/terms-conditions.html
Support: sinceregen.tech@gmail.com"""

DESC_DE="""ALLE IHRE VITALWERTE AN EINEM ORT

VitaFlow ist der umfassendste KI-Gesundheits-Tracker für das iPhone. Überwachen Sie Blutdruck, Herzfrequenz, Blutzucker, Blutsauerstoff (SpO2) und Schritte — in einer einzigen App.

HAUPTFUNKTIONEN

BLUTDRUCKMESSGERÄT
Systolisch, diastolisch und Puls. Klinische Farbkodierung. Wöchentliche und monatliche Trenddiagramme.

KI-REZEPTSCANNER
Richten Sie die Kamera auf ein Rezept. KI befüllt Ihre Medikamentenliste automatisch. Keine manuelle Eingabe. Kein anderer Gesundheits-Tracker hat diese Funktion.

HERZFREQUENZMESSUNG PER KAMERA
Legen Sie Ihren Finger auf die Rückkamera. Optische PPG-Technologie misst die Herzfrequenz in Sekunden. Kein Zusatzgerät.

BLUTZUCKER-TRACKER
mg/dL oder mmol/L. Mahlzeit-Kontextkennzeichnung. Klinische Farbkodierung.

BLUTSAUERSTOFF (SpO2)
Blutsauerstoffsättigung im Zeitverlauf.

SCHRITTZÄHLER
Echtzeit-Schrittzähler. Konfigurierbares Tagesziel. Apple Health-Integration.

MEDIKAMENTENERINNERUNGEN
Individueller Zeitplan. Schlummerfunktion. Funktioniert offline.

PDF- UND CSV-EXPORT
Direkt aus der App teilen.

APPLE WATCH
Messwerte vom Handgelenk. Automatische Synchronisation.

ICLOUD
Datensicherung in Ihrem iCloud-Konto. Wiederherstellung per Tastendruck.

DATENSCHUTZ
Alle Daten lokal. Kein Server. Keine Kontoregistrierung.

Datenschutz: https://sinceregen-legal.blogspot.com/p/privacy-policy.html
Nutzungsbedingungen: https://sinceregen-legal.blogspot.com/p/terms-conditions.html
Support: sinceregen.tech@gmail.com"""

DESC_ES="""TODOS TUS SIGNOS VITALES EN UN SOLO LUGAR

VitaFlow es el rastreador de salud con IA más completo para iPhone. Monitorea presión arterial, frecuencia cardíaca, glucosa en sangre, saturación de oxígeno (SpO2) y pasos diarios — todo en una app.

FUNCIONES PRINCIPALES

MONITOR DE PRESIÓN ARTERIAL
Sistólica, diastólica y pulso. Codificación de colores clínica. Gráficos de tendencia.

ESCÁNER DE RECETAS CON IA
Apunta tu cámara a cualquier receta. La IA rellena automáticamente tu lista de medicamentos. Sin entrada manual. Ningún otro rastreador de salud tiene esta función.

MONITOR DE FRECUENCIA CARDÍACA POR CÁMARA
Coloca tu dedo en la cámara trasera. Tecnología PPG óptica mide en segundos. Sin dispositivo adicional.

RASTREADOR DE GLUCOSA
mg/dL o mmol/L. Etiquetado de contexto de comida. Umbrales clínicos.

MONITOR DE OXÍGENO (SpO2)
Seguimiento de la saturación de oxígeno.

PODÓMETRO
Contador en tiempo real. Objetivo configurable. Apple Health integrado.

RECORDATORIO DE MEDICACIÓN
Horario personalizado. Posponer. Funciona sin conexión.

EXPORTACIÓN PDF Y CSV
Comparte directamente desde la app.

APPLE WATCH
Registra desde tu muñeca. Sincronización automática.

ICLOUD
Copia de seguridad en tu cuenta iCloud personal. Restauración con un toque.

PRIVACIDAD
Datos almacenados localmente. Sin servidor. Sin cuenta.

Privacidad: https://sinceregen-legal.blogspot.com/p/privacy-policy.html
Términos: https://sinceregen-legal.blogspot.com/p/terms-conditions.html
Soporte: sinceregen.tech@gmail.com"""

DESC_PT="""TODOS OS SEUS SINAIS VITAIS EM UM SÓ LUGAR

VitaFlow é o rastreador de saúde com IA mais completo para iPhone. Monitore pressão arterial, frequência cardíaca, glicose, saturação de oxigênio (SpO2) e passos diários — tudo em um app.

RECURSOS PRINCIPAIS

MONITOR DE PRESSÃO ARTERIAL
Sistólica, diastólica e pulso. Código de cores clínico. Gráficos de tendência semanais e mensais.

SCANNER DE RECEITAS COM IA
Aponte sua câmera para qualquer receita. A IA preenche automaticamente sua lista de medicamentos. Sem entrada manual. Nenhum outro rastreador de saúde tem esse recurso.

MONITOR DE FREQUÊNCIA CARDÍACA POR CÂMERA
Coloque o dedo na câmera traseira. Tecnologia PPG óptica mede em segundos. Sem dispositivo adicional.

RASTREADOR DE GLICOSE
mg/dL ou mmol/L. Marcação de contexto de refeição. Limites clínicos codificados por cor.

MONITOR DE OXIGÊNIO (SpO2)
Acompanhe a saturação de oxigênio ao longo do tempo.

PEDÔMETRO
Contador de passos em tempo real. Meta diária configurável. Integração Apple Health.

LEMBRETE DE MEDICAÇÃO
Horário personalizado. Soneca. Funciona offline.

EXPORTAÇÃO PDF E CSV
Compartilhe diretamente do app.

APPLE WATCH
Registre do pulso. Sincronização automática com iPhone.

ICLOUD
Backup na conta iCloud pessoal. Restauração em um toque.

PRIVACIDADE
Dados armazenados localmente. Sem servidor. Sem conta.

Privacidade: https://sinceregen-legal.blogspot.com/p/privacy-policy.html
Termos: https://sinceregen-legal.blogspot.com/p/terms-conditions.html
Suporte: sinceregen.tech@gmail.com"""

DESC_JA="""すべての健康データを一か所に

VitaFlow は iPhone 向け最も包括的な AI ヘルストラッカーです。血圧、心拍数、血糖値、血中酸素飽和度（SpO2）、歩数を一つのアプリで管理。

主な機能

血圧モニター
収縮期・拡張期・脈拍を記録。臨床閾値に基づくカラーコード表示。週次・月次トレンドチャート。

AI処方箋スキャナー
処方箋にカメラを向けるだけ。AIが薬名、用量、スケジュールを自動入力。手動入力不要。他の健康アプリにはない機能。

カメラ心拍数測定
指を後置カメラに当てるだけ。光学PPG技術が数秒で心拍数を計測。追加デバイス不要。

血糖値トラッカー
mg/dLとmmol/L対応。食事コンテキストタグ付き。臨床的な色分け表示。

血中酸素飽和度（SpO2）
経時的な血中酸素飽和度の追跡。

歩数計
リアルタイム計測、弧形ゲージ表示。日次目標設定可能。Apple Health完全対応。

薬リマインダー
カスタムスケジュール。スヌーズ1時間。オフラインで動作。

PDF・CSVエクスポート
アプリから直接共有。

Apple Watchコンパニオン
手首から記録。iPhoneと自動同期。

iCloudバックアップ
個人iCloudアカウントに全データ保存。ワンタップ復元。

プライバシー優先
全データをデバイスに保存。外部サーバーなし、アカウント不要。

プライバシー: https://sinceregen-legal.blogspot.com/p/privacy-policy.html
利用規約: https://sinceregen-legal.blogspot.com/p/terms-conditions.html
サポート: sinceregen.tech@gmail.com"""

DESC_KO="""모든 건강 수치를 한 곳에

VitaFlow는 iPhone을 위한 가장 완전한 AI 건강 트래커입니다. 혈압, 심박수, 혈당, 혈중 산소 포화도(SpO2), 일일 걸음 수를 하나의 앱으로 관리하세요.

주요 기능

혈압 모니터
수축기, 이완기, 맥박 기록. 임상 임계값 기반 색상 코딩. 주간/월간 추세 차트.

AI 처방전 스캐너
처방전에 카메라를 겨냥하면 됩니다. AI가 약 이름, 용량, 일정을 자동으로 입력합니다. 수동 입력 없음. 다른 건강 트래커에는 없는 기능.

카메라 심박수 측정
손가락을 후면 카메라에 대면 됩니다. 광학 PPG 기술이 몇 초 안에 심박수를 측정합니다. 추가 기기 불필요.

혈당 트래커
mg/dL 또는 mmol/L. 식사 맥락 태그. 임상 임계값 색상 코딩.

혈중 산소 포화도(SpO2)
시간에 따른 혈중 산소 포화도 추적.

만보기
실시간 계보기, 아크 게이지. 일일 목표 설정. Apple Health 완전 통합.

약 복용 알림
맞춤 일정. 1시간 다시 알림. 오프라인 작동.

PDF 및 CSV 내보내기
앱에서 직접 공유.

Apple Watch 동반
손목에서 기록. iPhone과 자동 동기화.

iCloud 백업
개인 iCloud 계정에 모든 데이터 백업. 원터치 복원.

개인 정보 보호
모든 데이터 로컬 저장. 외부 서버 없음, 계정 없음.

개인정보처리방침: https://sinceregen-legal.blogspot.com/p/privacy-policy.html
이용약관: https://sinceregen-legal.blogspot.com/p/terms-conditions.html
지원: sinceregen.tech@gmail.com"""

# ══════════════════════════════════════════════════════════════════════════════
# LOCALE DATA
# ══════════════════════════════════════════════════════════════════════════════
locs=[
  {"code":"en-US","lang":"English (U.S.) PRIMARY","us_sec":"PRIMARY","region":"United States","bg":CUS,
   "name":"VitaFlow: AI Health Tracker","nc":"27 ✓",
   "sub":"BP Heart Glucose SpO2 Monitor","sc":"29 ✓",
   "kw":"blood pressure,glucose,spo2,pedometer,prescription,icloud,apple watch,widget,oximeter,bmi,steps","kwc":"95 ✓",
   "promo":"Track blood pressure, heart rate, glucose, SpO2, and steps in one app. AI prescription scanner, Apple Watch companion, WidgetKit, iCloud backup. No server. No account.",
   "wn":WN_EN,"desc":DESC_EN},
  {"code":"ar","lang":"Arabic","us_sec":"YES ✓","region":"Saudi Arabia, UAE, Egypt, Gulf","bg":CAR,
   "name":"VitaFlow: متتبع الصحة AI","nc":"24 ✓",
   "sub":"ضغط الدم قلب جلوكوز SpO2","sc":"24 ✓",
   "kw":"ضغط الدم,مراقب,جلوكوز,أكسجين,خطوات,blood pressure,glucose,spo2,oximeter,prescription,pedometer","kwc":"~96 ✓",
   "promo":"تتبع ضغط الدم ومعدل ضربات القلب والجلوكوز وSpO2 والخطوات في تطبيق واحد. ماسح وصفات AI. مرافق Apple Watch. نسخ احتياطي iCloud.",
   "wn":WN_AR,"desc":DESC_AR},
  {"code":"zh-Hans","lang":"Chinese (Simplified)","us_sec":"YES ✓","region":"China, Singapore","bg":CZH,
   "name":"VitaFlow: AI健康追踪器","nc":"17 ✓",
   "sub":"血压心率血糖SpO2监测","sc":"17 ✓",
   "kw":"血压,心率,血糖,血氧,计步器,blood pressure,glucose,spo2,pedometer,prescription,apple watch,widget","kwc":"~88 ✓",
   "promo":"在一个应用中追踪血压、心率、血糖、SpO2和步数。AI处方扫描仪、Apple Watch伴侣、WidgetKit、iCloud备份。",
   "wn":WN_ZH_HANS,"desc":DESC_ZH_HANS},
  {"code":"zh-Hant","lang":"Chinese (Traditional)","us_sec":"YES ✓","region":"Taiwan, Hong Kong, Macau","bg":CZH,
   "name":"VitaFlow: AI健康追蹤器","nc":"17 ✓",
   "sub":"血壓心率血糖SpO2監測","sc":"17 ✓",
   "kw":"血壓,心率,血糖,血氧,計步器,blood pressure,glucose,spo2,pedometer,prescription,apple watch,widget","kwc":"~88 ✓",
   "promo":"在一個應用程式中追蹤血壓、心率、血糖、SpO2和步數。AI處方掃描器、Apple Watch伴侶、WidgetKit、iCloud備份。",
   "wn":WN_ZH_HANT,"desc":DESC_ZH_HANT},
  {"code":"en-AU","lang":"English (Australia)","us_sec":"No","region":"Australia, New Zealand","bg":CEN,
   "name":"VitaFlow: AI Health Tracker","nc":"27 ✓",
   "sub":"BP Heart Glucose SpO2 Monitor","sc":"29 ✓",
   "kw":"blood pressure,glucose,spo2,pedometer,prescription,icloud,apple watch,widget,oximeter,bmi,steps","kwc":"95 ✓",
   "promo":"Track blood pressure, heart rate, glucose, SpO2, and steps in one app. AI prescription scanner, Apple Watch, WidgetKit, iCloud backup. No server. No account.",
   "wn":WN_EN,"desc":DESC_EN},
  {"code":"en-CA","lang":"English (Canada)","us_sec":"No","region":"Canada","bg":CEN,
   "name":"VitaFlow: AI Health Tracker","nc":"27 ✓",
   "sub":"BP Heart Glucose SpO2 Monitor","sc":"29 ✓",
   "kw":"blood pressure,glucose,spo2,pedometer,prescription,icloud,apple watch,widget,oximeter,bmi,steps","kwc":"95 ✓",
   "promo":"Track blood pressure, heart rate, glucose, SpO2, and steps in one app. AI prescription scanner, Apple Watch, WidgetKit, iCloud backup. No server. No account.",
   "wn":WN_EN,"desc":DESC_EN},
  {"code":"en-GB","lang":"English (U.K.)","us_sec":"No","region":"UK + 80+ English countries","bg":CEN,
   "name":"VitaFlow: AI Health Tracker","nc":"27 ✓",
   "sub":"BP Heart Glucose SpO2 Monitor","sc":"29 ✓",
   "kw":"blood pressure,glucose,spo2,pedometer,prescription,icloud,apple watch,widget,oximeter,bmi,steps","kwc":"95 ✓",
   "promo":"Track blood pressure, heart rate, glucose, SpO2, and steps in one app. AI prescription scanner, Apple Watch, WidgetKit, iCloud backup. No server. No account.",
   "wn":WN_EN,"desc":DESC_EN},
  {"code":"fr-FR","lang":"French","us_sec":"YES ✓","region":"France, Belgium, Switzerland, Africa","bg":CFR,
   "name":"VitaFlow: Suivi Sante IA","nc":"24 ✓",
   "sub":"Tension Cardiaque Glucose SpO2","sc":"30 ✓",
   "kw":"pression artérielle,glycémie,oxymètre,podomètre,ordonnance,fréquence cardiaque,icloud,widget,tension,blood pressure","kwc":"~99 ✓",
   "promo":"Suivez tension artérielle, fréquence cardiaque, glycémie, SpO2 et pas dans une seule app. Scanner IA d'ordonnances, Apple Watch, widgets, iCloud.",
   "wn":WN_FR,"desc":DESC_FR},
  {"code":"de-DE","lang":"German","us_sec":"No","region":"Germany, Austria, Switzerland","bg":CDE,
   "name":"VitaFlow: KI Gesundheit Track","nc":"29 ✓",
   "sub":"Blutdruck Herzfrequenz Zucker","sc":"28 ✓",
   "kw":"blutdruck,blutzucker,herzfrequenz,spo2,blutdruck app,schrittzähler,rezept,icloud,apple watch,widget,oximeter","kwc":"~97 ✓",
   "promo":"Blutdruck, Herzfrequenz, Blutzucker, SpO2 und Schritte in einer App. KI-Rezeptscanner, Apple Watch, Widgets, iCloud. Kein Server, kein Konto.",
   "wn":WN_DE,"desc":DESC_DE},
  {"code":"ja-JP","lang":"Japanese","us_sec":"No","region":"Japan","bg":COT,
   "name":"VitaFlow: AI健康トラッカー","nc":"19 ✓",
   "sub":"血圧心拍血糖SpO2モニター","sc":"19 ✓",
   "kw":"血圧,心拍数,血糖値,血中酸素,万歩計,処方箋,icloud,apple watch,ウィジェット,血圧計,spo2","kwc":"~85 ✓",
   "promo":"血圧・心拍数・血糖値・SpO2・歩数を1つのアプリで管理。AI処方箋スキャナー、Apple Watchコンパニオン、WidgetKit、iCloudバックアップ。",
   "wn":WN_JA,"desc":DESC_JA},
  {"code":"ko-KR","lang":"Korean","us_sec":"YES ✓","region":"South Korea","bg":COT,
   "name":"VitaFlow: AI 건강 트래커","nc":"19 ✓",
   "sub":"혈압 심박 혈당 SpO2 모니터","sc":"22 ✓",
   "kw":"혈압,심박수,혈당,혈중산소,만보기,처방전,blood pressure,glucose,spo2,oximeter,apple watch","kwc":"~88 ✓",
   "promo":"혈압, 심박수, 혈당, SpO2, 걸음 수를 하나의 앱으로 추적. AI 처방전 스캐너, Apple Watch, 위젯, iCloud 백업.",
   "wn":WN_KO,"desc":DESC_KO},
  {"code":"pt-BR","lang":"Portuguese (Brazil)","us_sec":"YES ✓","region":"Brazil","bg":COT,
   "name":"VitaFlow: Monitor de Saude IA","nc":"29 ✓",
   "sub":"Pressao Cardiaca Glicose SpO2","sc":"28 ✓",
   "kw":"pressão arterial,glicose,spo2,frequência cardíaca,pedômetro,receita,icloud,apple watch,oxímetro,bmi,passos","kwc":"~99 ✓",
   "promo":"Monitore pressão arterial, frequência cardíaca, glicose, SpO2 e passos em um app. Scanner IA de receitas, Apple Watch, widgets, iCloud.",
   "wn":WN_PT,"desc":DESC_PT},
  {"code":"es-ES","lang":"Spanish (Spain)","us_sec":"No","region":"Spain","bg":COT,
   "name":"VitaFlow: Monitor de Salud IA","nc":"29 ✓",
   "sub":"Tension Cardiaco Glucosa SpO2","sc":"28 ✓",
   "kw":"presión arterial,glucosa,spo2,frecuencia cardíaca,podómetro,receta,icloud,apple watch,oxímetro,bmi,pasos","kwc":"~99 ✓",
   "promo":"Monitorea presión arterial, frecuencia cardíaca, glucosa, SpO2 y pasos en una app. Escáner IA de recetas, Apple Watch, widgets, iCloud.",
   "wn":WN_ES,"desc":DESC_ES},
  {"code":"es-MX","lang":"Spanish (Mexico) — ADD THIS","us_sec":"YES ✓","region":"Mexico + US Store bonus","bg":CMX,
   "name":"VitaFlow: Monitor de Salud IA","nc":"29 ✓",
   "sub":"Tension Cardiaco Glucosa SpO2","sc":"28 ✓",
   "kw":"blood pressure tracker,glucose monitor,heart rate,medication reminder,oximeter,spo2,pedometer,prescription scanner,health widget,icloud backup","kwc":"~99 ✓",
   "promo":"Monitorea presión arterial, frecuencia cardíaca, glucosa, SpO2 y pasos. Escáner IA de recetas, Apple Watch, widgets, iCloud. Sin servidor, sin cuenta.",
   "wn":WN_ES,"desc":DESC_ES},
]

locale_order=[(l["code"],l["lang"],l["bg"]) for l in locs]

screenshots=[
  {"num":"SS1","en_cap":"AI Health Tracker","en_sub":"BP Heart Glucose SpO2 Steps","locs":{
    "ar":("متتبع الصحة AI","ضغط القلب جلوكوز SpO2 خطوات"),
    "zh-Hans":("AI健康追踪器","血压心率血糖SpO2步数"),
    "zh-Hant":("AI健康追蹤器","血壓心率血糖SpO2步數"),
    "en-AU":("AI Health Tracker","BP Heart Glucose SpO2 Steps"),
    "en-CA":("AI Health Tracker","BP Heart Glucose SpO2 Steps"),
    "en-GB":("AI Health Tracker","BP Heart Glucose SpO2 Steps"),
    "fr-FR":("Tracker Sante IA","Tension Cardiaque Glucose SpO2 Pas"),
    "de-DE":("KI Gesundheit Tracker","Blutdruck Herzfrequenz Zucker SpO2"),
    "ja-JP":("AIヘルストラッカー","血圧心拍血糖SpO2歩数"),
    "ko-KR":("AI 건강 트래커","혈압 심박 혈당 SpO2 걸음"),
    "pt-BR":("Monitor de Saude IA","Pressao Cardiaca Glicose SpO2 Passos"),
    "es-ES":("Monitor Salud IA","Tension Cardiaco Glucosa SpO2 Pasos"),
    "es-MX":("Monitor Salud IA","Tension Cardiaco Glucosa SpO2 Pasos"),
  }},
  {"num":"SS2","en_cap":"Blood Pressure Monitor","en_sub":"Clinical Color Coded Readings","locs":{
    "ar":("مراقب ضغط الدم","قراءات مرمزة بالألوان السريرية"),
    "zh-Hans":("血压监测仪","临床色码读数"),
    "zh-Hant":("血壓監測儀","臨床色碼讀數"),
    "en-AU":("Blood Pressure Monitor","Clinical Color Coded Readings"),
    "en-CA":("Blood Pressure Monitor","Clinical Color Coded Readings"),
    "en-GB":("Blood Pressure Monitor","Clinical Colour Coded Readings"),
    "fr-FR":("Moniteur Pression Arterielle","Lectures Codées Couleurs Cliniques"),
    "de-DE":("Blutdruck Monitor","Klinisch Farbkodierte Werte"),
    "ja-JP":("血圧モニター","臨床カラーコード読取"),
    "ko-KR":("혈압 모니터","임상 색상 코딩 수치"),
    "pt-BR":("Monitor Pressao Arterial","Leituras Codigo de Cores Clinico"),
    "es-ES":("Monitor Presion Arterial","Lecturas Codificadas por Color"),
    "es-MX":("Monitor Presion Arterial","Lecturas Codificadas por Color"),
  }},
  {"num":"SS3","en_cap":"AI Prescription Scanner","en_sub":"Photo to Medication List Instantly","locs":{
    "ar":("ماسح الوصفات الطبية AI","من الصورة إلى قائمة الأدوية فوراً"),
    "zh-Hans":("AI处方扫描仪","拍照即刻生成药物清单"),
    "zh-Hant":("AI處方掃描器","拍照即刻生成藥物清單"),
    "en-AU":("AI Prescription Scanner","Photo to Medication List Instantly"),
    "en-CA":("AI Prescription Scanner","Photo to Medication List Instantly"),
    "en-GB":("AI Prescription Scanner","Photo to Medication List Instantly"),
    "fr-FR":("Scanner Ordonnances IA","Photo vers Liste Medicaments"),
    "de-DE":("KI Rezept Scanner","Foto zur Medikamentenliste"),
    "ja-JP":("AI処方箋スキャナー","写真で薬リストを即座に作成"),
    "ko-KR":("AI 처방전 스캐너","사진에서 즉시 약 목록 생성"),
    "pt-BR":("Scanner Receitas IA","Foto para Lista de Medicamentos"),
    "es-ES":("Escaner Recetas IA","Foto a Lista de Medicamentos"),
    "es-MX":("Escaner Recetas IA","Foto a Lista de Medicamentos"),
  }},
  {"num":"SS4","en_cap":"Camera Heart Rate Monitor","en_sub":"Optical PPG No Device Needed","locs":{
    "ar":("مراقب القلب بالكاميرا","PPG البصري لا جهاز مطلوب"),
    "zh-Hans":("摄像头心率监测器","光学PPG无需额外设备"),
    "zh-Hant":("攝影機心率監測器","光學PPG無需額外裝置"),
    "en-AU":("Camera Heart Rate Monitor","Optical PPG No Device Needed"),
    "en-CA":("Camera Heart Rate Monitor","Optical PPG No Device Needed"),
    "en-GB":("Camera Heart Rate Monitor","Optical PPG No Device Needed"),
    "fr-FR":("Moniteur Cardiaque Camera","PPG Optique Aucun Appareil"),
    "de-DE":("Kamera Herzfrequenz Monitor","Optisches PPG Kein Geraet Noetig"),
    "ja-JP":("カメラ心拍数モニター","光学PPG追加デバイス不要"),
    "ko-KR":("카메라 심박수 모니터","광학 PPG 추가 기기 불필요"),
    "pt-BR":("Monitor Cardiaco Camera","PPG Optico Sem Dispositivo"),
    "es-ES":("Monitor Cardiaco Camara","PPG Optico Sin Dispositivo"),
    "es-MX":("Monitor Cardiaco Camara","PPG Optico Sin Dispositivo"),
  }},
  {"num":"SS5","en_cap":"Blood Glucose Tracker","en_sub":"mg/dL mmol/L Meal Context","locs":{
    "ar":("متتبع جلوكوز الدم","mg/dL mmol/L سياق الوجبة"),
    "zh-Hans":("血糖追踪器","mg/dL mmol/L 餐食背景"),
    "zh-Hant":("血糖追蹤器","mg/dL mmol/L 餐食背景"),
    "en-AU":("Blood Glucose Tracker","mg/dL mmol/L Meal Context"),
    "en-CA":("Blood Glucose Tracker","mg/dL mmol/L Meal Context"),
    "en-GB":("Blood Glucose Tracker","mg/dL mmol/L Meal Context"),
    "fr-FR":("Suivi Glycemie","mg/dL mmol/L Contexte Repas"),
    "de-DE":("Blutzucker Tracker","mg/dL mmol/L Mahlzeitenkontext"),
    "ja-JP":("血糖値トラッカー","mg/dL mmol/L 食事コンテキスト"),
    "ko-KR":("혈당 트래커","mg/dL mmol/L 식사 맥락"),
    "pt-BR":("Rastreador Glicose","mg/dL mmol/L Contexto Refeicao"),
    "es-ES":("Rastreador Glucosa","mg/dL mmol/L Contexto Comida"),
    "es-MX":("Rastreador Glucosa","mg/dL mmol/L Contexto Comida"),
  }},
  {"num":"SS6","en_cap":"Step Counter Pedometer","en_sub":"Daily Goal Weekly Progress","locs":{
    "ar":("عداد الخطوات","الهدف اليومي التقدم الأسبوعي"),
    "zh-Hans":("计步器","每日目标 每周进度"),
    "zh-Hant":("計步器","每日目標 每週進度"),
    "en-AU":("Step Counter Pedometer","Daily Goal Weekly Progress"),
    "en-CA":("Step Counter Pedometer","Daily Goal Weekly Progress"),
    "en-GB":("Step Counter Pedometer","Daily Goal Weekly Progress"),
    "fr-FR":("Compteur de Pas Podometre","Objectif Quotidien Progres Hebdo"),
    "de-DE":("Schrittzaehler Pedometer","Tagesziel Woechentlicher Fortschritt"),
    "ja-JP":("歩数計ペドメーター","毎日の目標 週間進捗"),
    "ko-KR":("만보기 걸음 카운터","일일 목표 주간 진행"),
    "pt-BR":("Contador Passos Pedometro","Meta Diaria Progresso Semanal"),
    "es-ES":("Contador Pasos Podometro","Objetivo Diario Progreso Semanal"),
    "es-MX":("Contador Pasos Podometro","Objetivo Diario Progreso Semanal"),
  }},
  {"num":"SS7","en_cap":"Medication Reminder App","en_sub":"Custom Schedule Never Miss a Dose","locs":{
    "ar":("تطبيق تذكير الأدوية","جدول مخصص لا تفوت جرعة"),
    "zh-Hans":("药物提醒应用","自定义时间表 不错过剂量"),
    "zh-Hant":("藥物提醒應用","自訂時間表 不錯過劑量"),
    "en-AU":("Medication Reminder App","Custom Schedule Never Miss a Dose"),
    "en-CA":("Medication Reminder App","Custom Schedule Never Miss a Dose"),
    "en-GB":("Medication Reminder App","Custom Schedule Never Miss a Dose"),
    "fr-FR":("Rappel Medicaments App","Horaire Personnalise Ne Ratez Aucune Dose"),
    "de-DE":("Medikamentenerinnerung App","Zeitplan Keine Dosis Verpassen"),
    "ja-JP":("薬リマインダーアプリ","カスタムスケジュール 服薬を忘れない"),
    "ko-KR":("약 복용 알림 앱","맞춤 일정 복용 놓치지 않기"),
    "pt-BR":("Lembrete Medicacao App","Horario Personalizado Nao Perca Dose"),
    "es-ES":("Recordatorio Medicacion App","Horario Personalizado No Olvides Dosis"),
    "es-MX":("Recordatorio Medicacion App","Horario Personalizado No Olvides Dosis"),
  }},
  {"num":"SS8","en_cap":"Health Trends Charts","en_sub":"Weekly Monthly Vital History","locs":{
    "ar":("مخططات اتجاهات الصحة","السجل الأسبوعي والشهري"),
    "zh-Hans":("健康趋势图表","每周每月健康记录"),
    "zh-Hant":("健康趨勢圖表","每週每月健康記錄"),
    "en-AU":("Health Trends Charts","Weekly Monthly Vital History"),
    "en-CA":("Health Trends Charts","Weekly Monthly Vital History"),
    "en-GB":("Health Trends Charts","Weekly Monthly Vital History"),
    "fr-FR":("Graphiques Tendances Sante","Historique Vitaux Hebdo Mensuel"),
    "de-DE":("Gesundheits Trend Diagramme","Woechentliche Monatliche Historie"),
    "ja-JP":("健康トレンドチャート","週次月次バイタル履歴"),
    "ko-KR":("건강 추세 차트","주간 월간 건강 기록"),
    "pt-BR":("Graficos Tendencias Saude","Historico Semanal Mensal Vital"),
    "es-ES":("Graficos Tendencias Salud","Historial Semanal Mensual Vital"),
    "es-MX":("Graficos Tendencias Salud","Historial Semanal Mensual Vital"),
  }},
  {"num":"SS9","en_cap":"Apple Watch Health App","en_sub":"Log Vitals from Your Wrist","locs":{
    "ar":("تطبيق Apple Watch الصحي","سجّل المؤشرات من معصمك"),
    "zh-Hans":("Apple Watch健康应用","从手腕记录健康数据"),
    "zh-Hant":("Apple Watch健康應用","從手腕記錄健康數據"),
    "en-AU":("Apple Watch Health App","Log Vitals from Your Wrist"),
    "en-CA":("Apple Watch Health App","Log Vitals from Your Wrist"),
    "en-GB":("Apple Watch Health App","Log Vitals from Your Wrist"),
    "fr-FR":("Application Sante Apple Watch","Enregistrez Depuis le Poignet"),
    "de-DE":("Apple Watch Gesundheit App","Werte vom Handgelenk Erfassen"),
    "ja-JP":("Apple Watchヘルスアプリ","手首からバイタルを記録"),
    "ko-KR":("Apple Watch 건강 앱","손목에서 건강 수치 기록"),
    "pt-BR":("App Saude Apple Watch","Registre Sinais Vitais no Pulso"),
    "es-ES":("App Salud Apple Watch","Registra Signos Vitales del Muneca"),
    "es-MX":("App Salud Apple Watch","Registra Signos Vitales del Muneca"),
  }},
  {"num":"SS10","en_cap":"Health Widget Home Screen","en_sub":"All Vitals at a Glance","locs":{
    "ar":("ودجت الصحة الشاشة الرئيسية","جميع المؤشرات في لمحة"),
    "zh-Hans":("健康小组件主屏幕","所有健康数据一目了然"),
    "zh-Hant":("健康小工具主畫面","所有健康數據一目了然"),
    "en-AU":("Health Widget Home Screen","All Vitals at a Glance"),
    "en-CA":("Health Widget Home Screen","All Vitals at a Glance"),
    "en-GB":("Health Widget Home Screen","All Vitals at a Glance"),
    "fr-FR":("Widget Sante Ecran Accueil","Tous les Signes Vitaux en Un Coup"),
    "de-DE":("Gesundheit Widget Startbildschirm","Alle Werte auf einen Blick"),
    "ja-JP":("ヘルスウィジェットホーム画面","全バイタル一目で確認"),
    "ko-KR":("건강 위젯 홈 화면","모든 건강 수치 한눈에"),
    "pt-BR":("Widget Saude Tela Inicial","Todos os Sinais Vitais de Relance"),
    "es-ES":("Widget Salud Pantalla Inicio","Todos los Signos Vitales de Vistazo"),
    "es-MX":("Widget Salud Pantalla Inicio","Todos los Signos Vitales de Vistazo"),
  }},
]

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Overview
# ══════════════════════════════════════════════════════════════════════════════
ws0=wb.active; ws0.title="📋 Overview"
ws0.sheet_view.showGridLines=False
ws0.column_dimensions["A"].width=30; ws0.column_dimensions["B"].width=72

info=[
    ("VITAFLOW — ASO LOCALIZATION MASTER","",DARK,"FFFFFF",13,True),
    ("App Name: VitaFlow: AI Health Tracker  |  14 Locales  |  10 Screenshots  |  May 2026","",TEAL,"FFFFFF",10,True),
    ("","","FFFFFF","000000",10,False),
    ("SHEETS IN THIS FILE","","2D3748","FFFFFF",11,True),
    ("📋 Overview","This sheet — all rules, key decisions, colour guide.","EEF2FF","000000",10,False),
    ("🇺🇸 Metadata Table","Name, subtitle, keywords, promo text for all 14 locales with live char counts.","EEF2FF","000000",10,False),
    ("📸 Screenshot Captions","All 10 screenshots × 14 locales — ready for design handoff.","EEF2FF","000000",10,False),
    ("🆕 What's New","Full launch What's New text for all 14 locales.","EEF2FF","000000",10,False),
    ("📝 Descriptions","Full App Store descriptions for all 14 locales.","EEF2FF","000000",10,False),
    ("","","FFFFFF","000000",10,False),
    ("NAME DECISION","","2D3748","FFFFFF",11,True),
    ("Old name","VitalLog Pro — rejected (name conflict on App Store)","FFEBEE","A32D2D",10,False),
    ("Attempted","VitalScan: AI Health Tracker — rejected (name already in use)","FFEBEE","A32D2D",10,False),
    ("Final name","VitaFlow: AI Health Tracker (27 chars ✓) — unique, connects to FastFlow brand, full keyword value preserved","D4EDDA","000000",10,False),
    ("Why Arabic errors happened","The Arabic errors were a cascade from the name conflict — not independent issues. Changing the name to VitaFlow fixes all three Arabic errors automatically.","FFF3E0","000000",10,False),
    ("","","FFFFFF","000000",10,False),
    ("KEY RULES — READ BEFORE EDITING","","C53030","FFFFFF",11,True),
    ("App Name / Subtitle limit","Max 30 chars. NEVER use & or – (both count as 2 chars in Apple's algorithm). Keep under 28 to be completely safe. VitaFlow names are all 27-29 chars ✓","FFEBEE","000000",10,False),
    ("Keyword list","Max 100 chars. ZERO overlap with Name or Subtitle. Every char counts.","FFEBEE","000000",10,False),
    ("US Secondary locales","Arabic, zh-Hans, zh-Hant, French, Korean, PT-BR, ES-MX keyword fields are ALSO indexed for the US App Store — even though they are foreign-language localizations.","E8F5E9","000000",10,False),
    ("ES-MX keyword list","Spanish Mexico uses ENGLISH keywords only (not Spanish) in its keyword field. This maximises the US secondary bonus.","FFCCBC","C62828",10,False),
    ("Screenshot captions","2025 Apple algorithm: caption text is indexed as keywords. One keyword theme per screenshot. No & or – in any caption. Text at TOP of screenshot with high contrast.","FFF8E1","000000",10,False),
    ("IAP","Feature your IAP in App Store Connect immediately after launch — Apple boosts rankings for apps using platform features.","E8F5E9","000000",10,False),
    ("Rating prompt","Add SKStoreReviewManager.requestReview() triggered 3 seconds after first successful reading save. First 30 ratings are critical.","E8F5E9","000000",10,False),
    ("Screenshot order","Apple shows only first 3 screenshots in search results. SS3 (AI Prescription Scanner) is your #1 differentiator — keep it in positions 1-3.","FFF8E1","000000",10,False),
    ("","","FFFFFF","000000",10,False),
    ("ROW COLOUR GUIDE","","2D3748","FFFFFF",11,True),
    ("Green rows","English US primary","D4EDDA","000000",10,False),
    ("Orange rows","Arabic — key market for AI health apps","FFF3E0","000000",10,False),
    ("Yellow rows","Chinese Simplified and Traditional","FFF8E1","000000",10,False),
    ("Blue rows","English variants (AU, CA, GB) — same content as US","E3F2FD","000000",10,False),
    ("Purple rows","French","F3E5F5","000000",10,False),
    ("Indigo rows","German","E8EAF6","000000",10,False),
    ("Pink rows","Japanese, Korean, PT-BR, ES-ES","FBE9E7","000000",10,False),
    ("Peach rows","Spanish Mexico — ADD THIS LOCALE. Critical for US keyword bonus.","FFCCBC","C62828",10,False),
]

for i,(l,v,bg,fc,sz,bold) in enumerate(info,1):
    ca=ws0.cell(row=i,column=1,value=l); cb=ws0.cell(row=i,column=2,value=v)
    sc(ca,bg=bg,fc=fc,sz=sz,bold=bold); sc(cb,bg=bg,fc=fc,sz=sz,bold=False)
    if bold and not v: ws0.merge_cells(f"A{i}:B{i}")
    ws0.row_dimensions[i].height=18 if not v else 28
ws0.row_dimensions[1].height=22

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Metadata Table
# ══════════════════════════════════════════════════════════════════════════════
ws1=wb.create_sheet("🇺🇸 Metadata Table")
ws1.sheet_view.showGridLines=False
hdrs=["Language","Code","US Sec?","Region","App Name","Ch","Subtitle","Ch","Keywords (100 chars)","Ch","Promo Text (170 chars)"]
wds= [24,10,10,22,32,6,32,6,58,6,52]
for col,(h,w) in enumerate(zip(hdrs,wds),1):
    c=ws1.cell(row=1,column=col,value=h)
    sc(c,bg=DARK,fc="FFFFFF",sz=9,bold=True,ha="center")
    ws1.column_dimensions[get_column_letter(col)].width=w
ws1.row_dimensions[1].height=24
ws1.freeze_panes="E2"

for ri,loc in enumerate(locs,2):
    vals=[loc["lang"],loc["code"],loc["us_sec"],loc["region"],
          loc["name"],loc["nc"],loc["sub"],loc["sc"],
          loc["kw"],loc["kwc"],loc["promo"]]
    for ci,val in enumerate(vals,1):
        c=ws1.cell(row=ri,column=ci,value=val)
        if ci in (6,8,10):
            over="⚠" in str(val) or "TRIM" in str(val)
            sc(c,bg="FFCCCC" if over else "CCFFCC",fc="000000",sz=9,ha="center")
        else:
            sc(c,bg=loc["bg"],fc="000000",sz=9 if ci>4 else 10,bold=(ci<=2))
    ws1.row_dimensions[ri].height=44

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Screenshot Captions
# ══════════════════════════════════════════════════════════════════════════════
ws2=wb.create_sheet("📸 Screenshot Captions")
ws2.sheet_view.showGridLines=False

ss_hdrs=["Language","Region","SS1","SS2","SS3 ★","SS4","SS5","SS6","SS7","SS8","SS9","SS10"]
ss_wds= [22,22,28,26,28,26,24,24,26,24,26,26]
for col,(h,w) in enumerate(zip(ss_hdrs,ss_wds),1):
    bg=TEAL if "★" in h else DARK; fc="FFFFFF"
    c=ws2.cell(row=1,column=col,value=h)
    sc(c,bg=bg,fc=fc,sz=9,bold=True,ha="center")
    ws2.column_dimensions[get_column_letter(col)].width=w
ws2.row_dimensions[1].height=22

kw_row=["","",
    "(ai health tracker)","(blood pressure monitor)","(ai prescription scanner) ← KEY DIFFERENTIATOR",
    "(camera heart rate)","(blood glucose tracker)","(step counter pedometer)",
    "(medication reminder)","(health trends charts)","(apple watch health)","(health widget)"]
for col,t in enumerate(kw_row,1):
    c=ws2.cell(row=2,column=col,value=t)
    sc(c,bg=TEAL if col==5 else PURPLE,fc="FFFFFF",sz=8,ha="center",bold=False)
ws2.row_dimensions[2].height=16
ws2.freeze_panes="C3"

for ri,(code,lang,bg) in enumerate(locale_order,3):
    ws2.cell(row=ri,column=1,value=lang); sc(ws2.cell(row=ri,column=1),bg=bg,fc="000000",sz=10,bold=True)
    region=next((l["region"] for l in locs if l["code"]==code),"")
    ws2.cell(row=ri,column=2,value=region); sc(ws2.cell(row=ri,column=2),bg=bg,fc="555555",sz=9)
    for ci,ss in enumerate(screenshots,3):
        cap,sub=ss["locs"].get(code,(ss["en_cap"],ss["en_sub"])) if code!="en-US" else (ss["en_cap"],ss["en_sub"])
        c=ws2.cell(row=ri,column=ci,value=f"{cap}\n{sub}")
        sc(c,bg="FFFFFF",fc="000000",sz=9,ha="center")
    ws2.row_dimensions[ri].height=44

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — What's New
# ══════════════════════════════════════════════════════════════════════════════
ws3=wb.create_sheet("🆕 What's New")
ws3.sheet_view.showGridLines=False
ws3.column_dimensions["A"].width=26; ws3.column_dimensions["B"].width=12; ws3.column_dimensions["C"].width=85
for col,h in enumerate(["Language","Code","What's New — copy-paste into App Store Connect"],1):
    c=ws3.cell(row=1,column=col,value=h); sc(c,bg=DARK,fc="FFFFFF",sz=10,bold=True,ha="center")
ws3.row_dimensions[1].height=22; ws3.freeze_panes="C2"

for ri,(code,lang,bg) in enumerate(locale_order,2):
    ws3.cell(row=ri,column=1,value=lang); sc(ws3.cell(row=ri,column=1),bg=bg,fc="000000",sz=10,bold=True)
    ws3.cell(row=ri,column=2,value=code); sc(ws3.cell(row=ri,column=2),bg=bg,fc="555555",sz=9,ha="center")
    loc=next(l for l in locs if l["code"]==code)
    ws3.cell(row=ri,column=3,value=loc["wn"]); sc(ws3.cell(row=ri,column=3),bg=WHITE,fc="000000",sz=10)
    ws3.row_dimensions[ri].height=200

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — Descriptions
# ══════════════════════════════════════════════════════════════════════════════
ws4=wb.create_sheet("📝 Descriptions")
ws4.sheet_view.showGridLines=False
ws4.column_dimensions["A"].width=26; ws4.column_dimensions["B"].width=12; ws4.column_dimensions["C"].width=85
for col,h in enumerate(["Language","Code","Description — copy-paste into App Store Connect"],1):
    c=ws4.cell(row=1,column=col,value=h); sc(c,bg=DARK,fc="FFFFFF",sz=10,bold=True,ha="center")
ws4.row_dimensions[1].height=22; ws4.freeze_panes="C2"

for ri,(code,lang,bg) in enumerate(locale_order,2):
    ws4.cell(row=ri,column=1,value=lang); sc(ws4.cell(row=ri,column=1),bg=bg,fc="000000",sz=10,bold=True)
    ws4.cell(row=ri,column=2,value=code); sc(ws4.cell(row=ri,column=2),bg=bg,fc="555555",sz=9,ha="center")
    loc=next(l for l in locs if l["code"]==code)
    ws4.cell(row=ri,column=3,value=loc["desc"]); sc(ws4.cell(row=ri,column=3),bg=WHITE,fc="000000",sz=10)
    ws4.row_dimensions[ri].height=500

out="../VitalScan_ASO_Localizations.xlsx"
wb.save(out); print("Saved:",out)